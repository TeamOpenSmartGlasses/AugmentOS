from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
import gspread
import asyncio

import warnings
import time
from time import perf_counter, strftime, localtime
from functools import wraps
import os.path
import os

import openpyxl

# from constants import TIME_EVERYTHING

TIME_EVERYTHING = False
use_azure_openai = True

SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]
TOKEN_DIRECTORY = os.path.dirname(os.path.realpath(__file__))

WRITE_TO_LOCAL_SHEET = False
SAVE_THRESHOLD = 0.005
PRINT_ONLY = True  # If true, will print the timing information but not save it to the Google Sheet


def append_to_log_file(function_name, duration):
    with open("./helpers/log.txt", "a") as log_file:
        log_file.write(f"{function_name}: {duration:.6f} seconds\n")


def update_local_excel(function_name, time_called, duration):
    raise NotImplementedError("This function is not implemented yet")
    
    excel_file_path = os.path.join(TOKEN_DIRECTORY, "timings.xlsx")

    if os.path.exists(excel_file_path):
        workbook = openpyxl.load_workbook(excel_file_path)
    else:
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)

    if function_name in workbook.sheetnames:
        sheet = workbook[function_name]
    else:
        sheet = workbook.create_sheet(function_name)
        sheet.append(["Time Called", "Duration"])

    sheet.append([time_called, duration])

    workbook.save(excel_file_path)


def get_service():
    creds = None

    token_path = os.path.join(TOKEN_DIRECTORY, "token.json")
    if os.path.exists(token_path):
        creds = Credentials.from_authorized_user_file(token_path, SCOPES)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
            with open(token_path, "w") as token:
                token.write(creds.to_json())
        else:
            warnings.warn(
                "Google Sheets credentials are missing or invalid. Timing will not be "
                "recorded."
            )
            return None

    return gspread.authorize(creds)


def update_sheet_with_pandas(service, function_name, time_called, duration):
    pass
    # if use_azure_openai:
    #     function_name = "azure_" + function_name

    # if WRITE_TO_LOCAL_SHEET:
    #     update_local_excel(function_name, time_called, duration)

    # sh = service.open_by_key(time_everything_spreadsheet_id)
    # worksheet = sh.get_worksheet(1)

    # header_values = worksheet.row_values(1)
    # number_of_columns = len(header_values)

    # if function_name in header_values:
    #     function_index = header_values.index(function_name) + 1
    # else:
    #     function_index = number_of_columns + 2
    #     worksheet.update_cell(1, function_index, function_name)

    #     worksheet.update_cell(2, function_index, 'Time Called')
    #     worksheet.update_cell(2, function_index + 1, 'Duration')

    # next_empty_row = find_first_empty_row(worksheet, function_index)

    # timestamp_cell = gspread.utils.rowcol_to_a1(next_empty_row, function_index)
    # duration_cell = gspread.utils.rowcol_to_a1(next_empty_row, function_index + 1)

    # worksheet.update(f'{timestamp_cell}:{duration_cell}', [[time_called, duration]])


def find_first_empty_row(worksheet, col_index):
    col_values = worksheet.col_values(col_index)
    return len(col_values) + 1


def time_function():
    def decorator(function):
        if not TIME_EVERYTHING:
            return function

        time_called = strftime("%m-%d-%Y", localtime())

        try:
            @wraps(function)
            async def async_wrapper(*args, **kwargs):
                print(f"Timing function '{function.__name__}'")
                start_time = perf_counter()
                result = await function(*args, **kwargs)
                end_time = perf_counter()
                duration = end_time - start_time
                print(f"Function '{function.__name__}' took {duration:.6f} seconds")

                if PRINT_ONLY:
                    append_to_log_file(function.__name__, duration)
                    return result

                service = get_service()
                if not service:
                    warnings.warn(
                        "Unable to access Google Sheets. Timing will not be recorded."
                    )
                    return result

                function_name = function.__name__
                if function_name == "run_agent_async":
                    function_name += "_" + args[0].agent_name

                if duration > SAVE_THRESHOLD:
                    update_sheet_with_pandas(service, function_name, time_called, f"{duration:.6f} seconds")

                return result

            @wraps(function)
            def sync_wrapper(*args, **kwargs):
                print(f"Timing function '{function.__name__}'")
                start_time = perf_counter()
                result = function(*args, **kwargs)
                end_time = perf_counter()
                duration = end_time - start_time
                print(f"Function '{function.__name__}' took {duration:.6f} seconds")

                if PRINT_ONLY:
                    append_to_log_file(function.__name__, duration)
                    return result

                service = get_service()
                if not service:
                    warnings.warn(
                        "Unable to access Google Sheets. Timing will not be recorded."
                    )
                    return result

                function_name = function.__name__
                if function_name == "run_agent_async":
                    function_name += "_" + args[0].agent_name

                if duration > SAVE_THRESHOLD:
                    update_sheet_with_pandas(service, function.__name__, time_called, f"{duration:.6f} seconds")

                return result

            if asyncio.iscoroutinefunction(function):
                return async_wrapper
            else:
                return sync_wrapper
        
        except Exception as e:
            print(f"Error timing function '{function.__name__}': {e}")
            return function

    return decorator


if __name__ == "__main__":
    @time_function()
    def test_sync_function():
        time.sleep(0.1)


    @time_function()
    async def test_async_function():
        await asyncio.sleep(0.1)


    test_sync_function()
    asyncio.run(test_async_function())
