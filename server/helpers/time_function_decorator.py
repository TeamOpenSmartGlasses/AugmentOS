from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
import gspread
import asyncio

import warnings
import time
from time import perf_counter, strftime, localtime
from functools import wraps
import os.path
import os
import sys

import openpyxl
from openpyxl.utils import get_column_letter

WRITE_TO_LOCAL_SHEET = False

current_script_path = os.path.dirname(__file__)
base_dir = os.path.join(current_script_path, '..', '..')
server_dir = os.path.join(base_dir, 'server')
sys.path.append(server_dir)

from constants import TIME_EVERYTHING
from server_config import time_everything_spreadsheet_id, use_azure_openai

SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]

TOKEN_DIRECTORY = os.path.dirname(os.path.realpath(__file__))


def update_local_excel(function_name, time_called, duration):
    excel_file_path = os.path.join(TOKEN_DIRECTORY, "timings.xlsx")

    if os.path.exists(excel_file_path):
        workbook = openpyxl.load_workbook(excel_file_path)
    else:
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)

    if function_name in workbook.sheetnames:
        sheet = workbook[function_name]
    else:
        sheet = workbook.create_sheet(function_name)
        sheet.append(["Time Called", "Duration"])

    sheet.append([time_called, duration])

    workbook.save(excel_file_path)


def get_service():
    creds = None

    token_path = os.path.join(TOKEN_DIRECTORY, "token.json")
    if os.path.exists(token_path):
        creds = Credentials.from_authorized_user_file(token_path, SCOPES)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
            with open(token_path, "w") as token:
                token.write(creds.to_json())
        else:
            warnings.warn(
                "Google Sheets credentials are missing or invalid. Timing will not be "
                "recorded."
            )
            return None

    return gspread.authorize(creds)


def update_sheet_with_pandas(service, function_name, time_called, duration):
    if use_azure_openai:
        function_name = "azure_" + function_name

    if WRITE_TO_LOCAL_SHEET:
        update_local_excel(function_name, time_called, duration)

    sh = service.open_by_key(time_everything_spreadsheet_id)
    worksheet = sh.get_worksheet(0)

    header_values = worksheet.row_values(1)
    number_of_columns = len(header_values)

    if function_name in header_values:
        function_index = header_values.index(function_name) + 1
    else:
        function_index = number_of_columns + 2
        worksheet.update_cell(1, function_index, function_name)

    next_empty_row = find_first_empty_row(worksheet, function_index)

    timestamp_cell = gspread.utils.rowcol_to_a1(next_empty_row, function_index)
    duration_cell = gspread.utils.rowcol_to_a1(next_empty_row, function_index + 1)

    worksheet.update(f'{timestamp_cell}:{duration_cell}', [[time_called, duration]])


def find_first_empty_row(worksheet, col_index):
    col_values = worksheet.col_values(col_index)
    return len(col_values) + 1


def time_function():
    def decorator(function):
        if not TIME_EVERYTHING:
            return function

        @wraps(function)
        async def async_wrapper(*args, **kwargs):
            print(f"Timing function '{function.__name__}'")
            service = get_service()
            if not service:
                warnings.warn(
                    "Unable to access Google Sheets. Timing will not be recorded."
                )
                return function(*args, **kwargs)

            start_time = perf_counter()
            result = await function(*args, **kwargs)
            end_time = perf_counter()
            duration = end_time - start_time
            print(f"Function '{function.__name__}' took {duration:.6f} seconds")

            time_called = strftime("%Y-%m-%d %H:%M:%S", localtime())

            update_sheet_with_pandas(service, function.__name__, time_called,
                f"{duration:.6f} seconds"
            )

            return result

        @wraps(function)
        def sync_wrapper(*args, **kwargs):
            print(f"Timing function '{function.__name__}'")
            service = get_service()
            if not service:
                warnings.warn(
                    "Unable to access Google Sheets. Timing will not be recorded."
                )
                return function(*args, **kwargs)

            start_time = perf_counter()
            result = function(*args, **kwargs)
            end_time = perf_counter()
            duration = end_time - start_time
            print(f"Function '{function.__name__}' took {duration:.6f} seconds")

            time_called = strftime("%Y-%m-%d %H:%M:%S", localtime())

            update_sheet_with_pandas(service, function.__name__, time_called,
                f"{duration:.6f} seconds"
            )

            return result

        if asyncio.iscoroutinefunction(function):
            return async_wrapper
        else:
            return sync_wrapper

    return decorator


if __name__ == "__main__":

    @time_function()
    def some_function():
        time.sleep(0.1)


    some_function()
